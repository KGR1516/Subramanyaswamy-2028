"""
build_report.py
----------------
Turns one run_daily.py result dict into a formula-driven Excel workbook.

House style: Arial · blue text = value produced by the analysis pipeline
(price, verdict, confidence — the report's inputs) · black text = an Excel
formula computed from cells in this workbook · green text = a formula that
links to another sheet. "Fired" is a real formula (Verdict + Confidence vs.
the threshold cell), not a hardcoded flag, so it recalculates if you change
the threshold on the Summary tab.

Sheet 2 ("All Verdicts") carries both the original verdict columns and the
technical-analysis columns (RSI-14, MACD signal, SMA50/200 cross, ADX-14,
Volume Breakout) and fundamental-analysis columns (P/E, ROE, revenue
growth, debt/equity, 5-year revenue & net-income CAGR, Free/Operating/Net
Cash Flow, EBIT, EBITDA, Cash Conversion Cycle) — see data_sources.py /
scoring.py for what feeds them. New columns are always appended at the end
of VERDICT_HEADERS, never inserted in the middle — several Summary-tab and
conditional-formatting formulas hardcode column letters (L/N/O/Q/R) that
would silently break if earlier columns shifted.

Sheet 3 ("5Y Fundamentals") is the multi-year detail behind those CAGR
columns: revenue and net income for every year actually available (Yahoo's
free feed usually caps annual statements below 5 — that shortfall is
labeled, never padded) per shortlisted stock, with CAGR computed as real
Excel formulas from the displayed year cells (not baked-in Python floats).

Can be run standalone against a saved result for testing:
    python build_report.py path/to/result.json out.xlsx
"""
import sys
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1A4D5C")
HEADER_FONT = Font(name=FONT, size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, size=16, bold=True, color="1A4D5C")
SUB_FONT = Font(name=FONT, size=10, italic=True, color="6B675F")
INPUT_FONT = Font(name=FONT, size=10, color="0000FF")     # blue = pipeline output (input to this sheet)
FORMULA_FONT = Font(name=FONT, size=10, color="000000")   # black = in-sheet formula
LINK_FONT = Font(name=FONT, size=10, color="008000")      # green = cross-sheet link
NA_FONT = Font(name=FONT, size=10, color="9A958A")        # gray = no data available
THIN = Side(style="thin", color="D8D3C7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

VERDICT_HEADERS = [
    "Symbol", "Name", "Cap Segment", "Verdict", "Confidence",
    "Winner", "Price (₹)", "Day Change %", "Key Catalyst",
    "Rationale", "Engine", "Fired",
    # --- technical analysis -------------------------------------------------
    "RSI (14)", "MACD Signal", "SMA50/200 Cross", "ADX (14)",
    # --- fundamental analysis ------------------------------------------------
    "P/E (Trailing)", "ROE %", "Revenue Gr. YoY %", "Debt/Equity",
    "Revenue CAGR (5Y) %", "Net Income CAGR (5Y) %", "Fund. Years",
    # --- volume breakout + cash flow (appended, not inserted, so the
    # hardcoded Summary-tab / conditional-formatting column letters above
    # — L/N/O/Q/R — never shift) ----------------------------------------
    "Volume Breakout", "Free Cash Flow", "FCF Yield %",
    "Operating Cash Flow", "Net Cash Flow", "EBIT", "EBITDA", "CCC (Days)",
]

FUND5Y_HEADERS = [
    "Symbol", "Name", "Fiscal Years",
    "Revenue Y-4", "Revenue Y-3", "Revenue Y-2", "Revenue Y-1", "Revenue (Latest)",
    "Net Income Y-4", "Net Income Y-3", "Net Income Y-2", "Net Income Y-1", "Net Income (Latest)",
    "Revenue CAGR %", "Net Income CAGR %", "Net Margin % (Latest Yr)", "Years Available (of 5)",
]


def _blank_or(v, font_if_value):
    """Return (value, font) — blank cell with a gray N/A note if v is None."""
    if v is None:
        return "—", NA_FONT
    return v, font_if_value


def build(result, output_path="daily_digest.xlsx"):
    verdicts = result["verdicts"]
    fundamentals_5y = result.get("fundamentals_5y", [])
    last_row = len(verdicts) + 1  # header is row 1

    # the "BUY confidence" helper column lives one past the last visible
    # header — computed once, up front, so the Summary tab's cross-sheet
    # formulas and the All Verdicts sheet that actually writes it agree on
    # its position (it moves whenever VERDICT_HEADERS grows).
    helper_col = len(VERDICT_HEADERS) + 1
    helper_letter = get_column_letter(helper_col)

    wb = Workbook()

    # -----------------------------------------------------------------
    # Sheet 1: Summary
    # -----------------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24

    ws["A1"] = "Bourse — Daily Analysis Digest"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {result['timestamp']}  ·  engine: {result['engine']}  ·  mode: {result['mode']}"
    ws["A2"].font = SUB_FONT

    ws["A3"] = ("Live NSE data pulled via yfinance and run through the same Scout -> Technician -> "
                "Fundamentalist -> Newsdesk -> Bull/Bear -> Judge pipeline as the local dashboard's "
                "Live mode — run here on GitHub Actions, which (unlike a locked-down sandbox) has "
                "normal internet access. Technicals now include RSI/MACD/Bollinger/SMA50-200/ADX; "
                "fundamentals include P/E, ROE, margins, leverage, and a multi-year revenue/earnings "
                "trend (see the 'All Verdicts' and '5Y Fundamentals' tabs).")
    ws["A3"].font = Font(name=FONT, size=9, italic=True, color="6B675F")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A3:D3")
    ws.row_dimensions[3].height = 56

    ws["A5"] = "Confidence threshold (BUY fires at ≥)"
    ws["A5"].font = Font(name=FONT, size=10, bold=True)
    thr_cell = ws.cell(row=5, column=2, value=result["confidence_threshold"])
    thr_cell.font = INPUT_FONT
    thr_cell.fill = PatternFill("solid", fgColor="FFFFCC")
    THRESHOLD_REF = "Summary!$B$5"

    rows = [
        ("Universe scanned", result["universe"]),
        ("Shortlisted / in debate", result["in_debate"]),
        ("BUY signals fired", f"=COUNTIF('All Verdicts'!L2:L{last_row},\"YES\")"),
        ("Avg confidence (all)", f"=IFERROR(ROUND(AVERAGE('All Verdicts'!E2:E{last_row}),2),\"—\")"),
        ("Top BUY confidence", f"=IFERROR(MAX('All Verdicts'!{helper_letter}2:{helper_letter}{last_row}),\"—\")"),
        ("Top pick", f"=IFERROR(INDEX('All Verdicts'!A2:A{last_row},"
                      f"MATCH(MAX('All Verdicts'!{helper_letter}2:{helper_letter}{last_row}),"
                      f"'All Verdicts'!{helper_letter}2:{helper_letter}{last_row},0)),\"—\")"),
        # P/E averaged only over positive values — a negative P/E (an
        # unprofitable company) isn't a valuation multiple you can blend
        # into an average with profitable ones without distorting it.
        ("Avg P/E (where available, profitable only)",
         f"=IFERROR(ROUND(AVERAGEIF('All Verdicts'!Q2:Q{last_row},\">0\"),1),\"—\")"),
        ("Avg ROE % (where available)", f"=IFERROR(ROUND(AVERAGEIF('All Verdicts'!R2:R{last_row},\"<>—\"),1),\"—\")"),
    ]
    row = 7
    for label, val in rows:
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=True)
        c = ws.cell(row=row, column=2, value=val)
        c.font = INPUT_FONT if isinstance(val, (int, float)) else LINK_FONT
        row += 1

    note_row = row + 1
    ws.cell(row=note_row, column=1,
            value=("Full technical + fundamental breakdown on 'All Verdicts'; multi-year "
                   "revenue/earnings detail on '5Y Fundamentals'. Same analysis engine as "
                   "the local dashboard — not a lighter scrape — because GitHub Actions can "
                   "reach yfinance directly."))
    ws.cell(row=note_row, column=1).font = Font(name=FONT, size=9, italic=True, color="6B675F")
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    ws.row_dimensions[note_row].height = 32

    legend_row = note_row + 2
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(name=FONT, size=9, bold=True)
    ws.cell(row=legend_row + 1, column=1, value="Blue = pipeline output (this report's input)").font = INPUT_FONT
    ws.cell(row=legend_row + 2, column=1,
            value="Black / green = formula (recalculates; green links to another sheet)").font = LINK_FONT
    ws.cell(row=legend_row + 3, column=1, value="Gray “—” = not available for this stock/period").font = NA_FONT
    ws.cell(row=legend_row + 4, column=1,
            value="Yellow fill = edit this to change the BUY threshold used by the Fired column").font = Font(
        name=FONT, size=9, italic=True, color="6B675F")

    # -----------------------------------------------------------------
    # Sheet 2: All Verdicts
    # -----------------------------------------------------------------
    ws2 = wb.create_sheet("All Verdicts")
    for col, h in enumerate(VERDICT_HEADERS, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws2.row_dimensions[1].height = 34
    ws2.freeze_panes = "A2"

    widths = [12, 20, 12, 10, 11, 9, 12, 12, 26, 34, 13, 8,
              10, 13, 15, 10,
              13, 9, 15, 12, 16, 18, 11,
              14, 15, 11, 16, 14, 14, 14, 12]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    for i, v in enumerate(verdicts, start=2):
        ws2.cell(row=i, column=1, value=v["symbol"]).font = Font(name=FONT, size=10, bold=True)
        ws2.cell(row=i, column=2, value=v["name"]).font = INPUT_FONT
        ws2.cell(row=i, column=3, value=v["cap"].capitalize() if v["cap"] else v["cap"]).font = INPUT_FONT
        ws2.cell(row=i, column=4, value=v["verdict"]).font = INPUT_FONT

        cc = ws2.cell(row=i, column=5, value=v["confidence"])
        cc.font = INPUT_FONT

        ws2.cell(row=i, column=6, value=v["winner"]).font = INPUT_FONT

        pc = ws2.cell(row=i, column=7, value=v["price"])
        pc.font = INPUT_FONT
        pc.number_format = "#,##0.00"

        dc = ws2.cell(row=i, column=8, value=v["day_change_pct"])
        dc.font = INPUT_FONT
        dc.number_format = "0.00"

        ws2.cell(row=i, column=9, value=v["key_catalyst"]).font = INPUT_FONT
        rc = ws2.cell(row=i, column=10, value=v["rationale"])
        rc.font = INPUT_FONT
        rc.alignment = Alignment(wrap_text=True, vertical="top")

        ws2.cell(row=i, column=11, value=v["engine"]).font = Font(name=FONT, size=9, color="6B675F")

        # Fired: real formula against the Summary threshold cell, not the
        # Python-computed boolean, so editing the threshold recalculates it.
        fired_formula = f'=IF(AND(D{i}="BUY",E{i}>={THRESHOLD_REF}),"YES","-")'
        fc = ws2.cell(row=i, column=12, value=fired_formula)
        fc.font = LINK_FONT
        fc.alignment = Alignment(horizontal="center")

        # --- technical analysis columns (13-16) -----------------------------
        val, font = _blank_or(v.get("rsi14"), INPUT_FONT)
        ws2.cell(row=i, column=13, value=val).font = font

        macd = v.get("macd_bullish")
        macd_label = "Bullish" if macd is True else ("Bearish" if macd is False else "—")
        ws2.cell(row=i, column=14, value=macd_label).font = INPUT_FONT if macd is not None else NA_FONT

        cross = v.get("sma_cross")
        cross_label = {"golden": "Golden", "death": "Death", "flat": "Flat"}.get(cross, "—")
        ws2.cell(row=i, column=15, value=cross_label).font = INPUT_FONT if cross else NA_FONT

        val, font = _blank_or(v.get("adx14"), INPUT_FONT)
        ws2.cell(row=i, column=16, value=val).font = font

        # --- fundamental analysis columns (17-23) ---------------------------
        val, font = _blank_or(v.get("pe_trailing"), INPUT_FONT)
        ws2.cell(row=i, column=17, value=val).font = font

        val, font = _blank_or(v.get("roe_pct"), INPUT_FONT)
        ws2.cell(row=i, column=18, value=val).font = font

        val, font = _blank_or(v.get("revenue_growth_yoy_pct"), INPUT_FONT)
        ws2.cell(row=i, column=19, value=val).font = font

        val, font = _blank_or(v.get("debt_to_equity"), INPUT_FONT)
        ws2.cell(row=i, column=20, value=val).font = font

        val, font = _blank_or(v.get("revenue_cagr_pct"), INPUT_FONT)
        ws2.cell(row=i, column=21, value=val).font = font

        val, font = _blank_or(v.get("net_income_cagr_pct"), INPUT_FONT)
        ws2.cell(row=i, column=22, value=val).font = font

        val, font = _blank_or(v.get("fundamentals_years"), INPUT_FONT)
        ws2.cell(row=i, column=23, value=val).font = font

        # --- volume breakout + cash flow columns (24-31) ---------------------
        vb = v.get("volume_breakout")
        vb_label = "Yes" if vb is True else ("No" if vb is False else "—")
        ws2.cell(row=i, column=24, value=vb_label).font = INPUT_FONT if vb is not None else NA_FONT

        val, font = _blank_or(v.get("free_cash_flow"), INPUT_FONT)
        fcf_c = ws2.cell(row=i, column=25, value=val)
        fcf_c.font = font
        if val != "—":
            fcf_c.number_format = "#,##0"

        val, font = _blank_or(v.get("fcf_yield_pct"), INPUT_FONT)
        ws2.cell(row=i, column=26, value=val).font = font

        val, font = _blank_or(v.get("operating_cash_flow"), INPUT_FONT)
        ocf_c = ws2.cell(row=i, column=27, value=val)
        ocf_c.font = font
        if val != "—":
            ocf_c.number_format = "#,##0"

        val, font = _blank_or(v.get("net_cash_flow"), INPUT_FONT)
        ncf_c = ws2.cell(row=i, column=28, value=val)
        ncf_c.font = font
        if val != "—":
            ncf_c.number_format = "#,##0"

        val, font = _blank_or(v.get("ebit"), INPUT_FONT)
        ebit_c = ws2.cell(row=i, column=29, value=val)
        ebit_c.font = font
        if val != "—":
            ebit_c.number_format = "#,##0"

        val, font = _blank_or(v.get("ebitda"), INPUT_FONT)
        ebitda_c = ws2.cell(row=i, column=30, value=val)
        ebitda_c.font = font
        if val != "—":
            ebitda_c.number_format = "#,##0"

        val, font = _blank_or(v.get("cash_conversion_cycle_days"), INPUT_FONT)
        ws2.cell(row=i, column=31, value=val).font = font

        for col in range(1, len(VERDICT_HEADERS) + 1):
            ws2.cell(row=i, column=col).border = BORDER

    # helper column M (BUY-only confidence) — kept at column 13 to match the
    # Summary tab's existing formulas ('All Verdicts'!M2:M...), so it's
    # written into its own dedicated column *after* the visible table rather
    # than shifting every downstream reference.
    helper_col = len(VERDICT_HEADERS) + 1
    helper_letter = get_column_letter(helper_col)
    ws2.cell(row=1, column=helper_col, value="BUY Confidence (helper)").font = HEADER_FONT
    ws2.cell(row=1, column=helper_col).fill = HEADER_FILL
    ws2.column_dimensions[helper_letter].width = 20
    for i in range(2, last_row + 1):
        mc = ws2.cell(row=i, column=helper_col, value=f'=IF(D{i}="BUY",E{i},"")')
        mc.font = FORMULA_FONT

    ws2.auto_filter.ref = f"A1:{helper_letter}{last_row}"

    # conditional formatting
    green = PatternFill("solid", fgColor="E8F4EE")
    red = PatternFill("solid", fgColor="FBE9E7")
    gray = PatternFill("solid", fgColor="EFECE6")
    ws2.conditional_formatting.add(f"D2:D{last_row}",
        CellIsRule(operator="equal", formula=['"BUY"'], fill=green))
    ws2.conditional_formatting.add(f"D2:D{last_row}",
        CellIsRule(operator="equal", formula=['"AVOID"'], fill=red))
    ws2.conditional_formatting.add(f"D2:D{last_row}",
        CellIsRule(operator="equal", formula=['"WATCH"'], fill=gray))
    ws2.conditional_formatting.add(f"H2:H{last_row}", FormulaRule(formula=["H2>0"], fill=green))
    ws2.conditional_formatting.add(f"H2:H{last_row}", FormulaRule(formula=["H2<0"], fill=red))
    ws2.conditional_formatting.add(f"L2:L{last_row}",
        CellIsRule(operator="equal", formula=['"YES"'], fill=green))
    ws2.conditional_formatting.add(f"N2:N{last_row}",
        CellIsRule(operator="equal", formula=['"Bullish"'], fill=green))
    ws2.conditional_formatting.add(f"N2:N{last_row}",
        CellIsRule(operator="equal", formula=['"Bearish"'], fill=red))
    ws2.conditional_formatting.add(f"O2:O{last_row}",
        CellIsRule(operator="equal", formula=['"Golden"'], fill=green))
    ws2.conditional_formatting.add(f"O2:O{last_row}",
        CellIsRule(operator="equal", formula=['"Death"'], fill=red))

    # volume breakout + cash flow columns (24-31 -> X, Y, Z, AA, AB, AC, AD, AE)
    vb_letter = get_column_letter(24)
    fcf_letter = get_column_letter(25)
    ocf_letter = get_column_letter(27)
    ncf_letter = get_column_letter(28)
    ebit_letter = get_column_letter(29)
    ebitda_letter = get_column_letter(30)
    ws2.conditional_formatting.add(f"{vb_letter}2:{vb_letter}{last_row}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=green))
    for letter in (fcf_letter, ocf_letter, ncf_letter, ebit_letter, ebitda_letter):
        rng = f"{letter}2:{letter}{last_row}"
        ws2.conditional_formatting.add(rng, FormulaRule(formula=[f"{letter}2>0"], fill=green))
        ws2.conditional_formatting.add(rng, FormulaRule(formula=[f"{letter}2<0"], fill=red))

    # -----------------------------------------------------------------
    # Sheet 3: 5Y Fundamentals
    # -----------------------------------------------------------------
    ws3 = wb.create_sheet("5Y Fundamentals")
    for col, h in enumerate(FUND5Y_HEADERS, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws3.row_dimensions[1].height = 34
    ws3.freeze_panes = "A2"

    f5y_widths = [12, 20, 22, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 14, 16, 20, 16]
    for i, w in enumerate(f5y_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3.cell(row=1, column=1)  # anchor before note
    note = ws3.cell(row=1, column=1)
    if fundamentals_5y:
        last3 = len(fundamentals_5y) + 1
        for i, s in enumerate(fundamentals_5y, start=2):
            years = s.get("years", []) or []           # chronological, oldest first
            rev_by_year = s.get("revenue_by_year", {}) or {}
            ni_by_year = s.get("net_income_by_year", {}) or {}
            n = min(len(years), 5)

            ws3.cell(row=i, column=1, value=s["symbol"]).font = Font(name=FONT, size=10, bold=True)
            ws3.cell(row=i, column=2, value=s.get("name", s["symbol"])).font = INPUT_FONT
            ws3.cell(row=i, column=3, value=", ".join(years) if years else "—").font = INPUT_FONT

            # right-align into the 5-year block: most recent always lands in
            # the rightmost column (col 8 for revenue, col 13 for net income)
            # so a stock with only 4 years of Yahoo data leaves Y-4 blank
            # instead of misaligning every stock's earliest year together.
            rev_start_col = 4 + (5 - n)
            ni_start_col = 9 + (5 - n)
            for j, y in enumerate(years[-5:]):
                rv = rev_by_year.get(y)
                c = ws3.cell(row=i, column=rev_start_col + j, value=rv if rv is not None else "—")
                c.font = INPUT_FONT if rv is not None else NA_FONT
                c.number_format = "#,##0"
                nv = ni_by_year.get(y)
                c2 = ws3.cell(row=i, column=ni_start_col + j, value=nv if nv is not None else "—")
                c2.font = INPUT_FONT if nv is not None else NA_FONT
                c2.number_format = "#,##0"

            rev_range = f"D{i}:H{i}"
            ni_range = f"I{i}:M{i}"
            # CAGR as a real formula off the displayed year cells: oldest
            # populated cell is found via INDEX/COUNT (right-aligned layout
            # means "oldest" is always (6 - count) cells in from the left),
            # newest is always the rightmost cell of the block.
            rev_cagr = (f'=IFERROR(IF(COUNT({rev_range})>=2,'
                        f'(H{i}/INDEX({rev_range},6-COUNT({rev_range})))^'
                        f'(1/(COUNT({rev_range})-1))-1,"—"),"—")')
            ni_cagr = (f'=IFERROR(IF(COUNT({ni_range})>=2,'
                       f'(M{i}/INDEX({ni_range},6-COUNT({ni_range})))^'
                       f'(1/(COUNT({ni_range})-1))-1,"—"),"—")')
            margin_latest = f'=IFERROR(M{i}/H{i},"—")'

            c = ws3.cell(row=i, column=14, value=rev_cagr)
            c.font = FORMULA_FONT
            c.number_format = "0.0%"
            c = ws3.cell(row=i, column=15, value=ni_cagr)
            c.font = FORMULA_FONT
            c.number_format = "0.0%"
            c = ws3.cell(row=i, column=16, value=margin_latest)
            c.font = FORMULA_FONT
            c.number_format = "0.0%"

            yc = ws3.cell(row=i, column=17, value=f"=COUNT({rev_range})")
            yc.font = FORMULA_FONT

            for col in range(1, len(FUND5Y_HEADERS) + 1):
                ws3.cell(row=i, column=col).border = BORDER

        ws3.auto_filter.ref = f"A1:Q{last3}"
        note_row3 = last3 + 2
    else:
        note_row3 = 3

    note = ws3.cell(row=note_row3, column=1,
        value=("Revenue and Net Income figures are as reported in each company's annual "
               "financial statements (source: yfinance / Yahoo Finance). Yahoo's free feed "
               "typically exposes up to ~4 years of annual statements, not a full 5 — where a "
               "stock shows fewer than 5 populated year columns, that is a genuine data "
               "availability limit, not an omission. CAGR and latest-year net margin are live "
               "formulas computed from the year columns in this sheet."))
    note.font = Font(name=FONT, size=9, italic=True, color="6B675F")
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws3.merge_cells(start_row=note_row3, start_column=1, end_row=note_row3, end_column=8)
    ws3.row_dimensions[note_row3].height = 48

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    in_path = sys.argv[1] if len(sys.argv) > 1 else "result.json"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "daily_digest.xlsx"
    with open(in_path) as f:
        result = json.load(f)
    build(result, out_path)
    print(f"saved {out_path}")
